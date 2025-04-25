"""
Module that defines the MCDM project management service.

This service is responsible for the persistence, retrieval, and general management of projects.
"""
from typing import Dict, List, Any, Optional, Tuple, Set
import os
import json
import pandas as pd
import csv

import csv
from domain.entities.alternative import Alternative
from domain.entities.criteria import Criteria, OptimizationType, ScaleType
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from domain.entities.project import Project
from domain.repositories.project_repository import ProjectRepository
from application.validators.project_validator import ProjectValidator
from utils.exceptions import ServiceError, ValidationError, RepositoryError
from reportlab.lib.pagesizes import letter
from reportlab.lib import colors
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet


class ProjectService:
    
    def __init__(self, repository: ProjectRepository):
        self._repository = repository
    
    def save_project(self, project: Project) -> Project:
        try:
            is_valid, errors = ProjectValidator.validate_project(project)
            if not is_valid:
                raise ValidationError(
                    message="The project is not valid",
                    errors=errors
                )
            saved_project = self._repository.save(project)
            
            return saved_project
            
        except ValidationError as e:
            raise ServiceError(
                message=f"Validation error when saving project: {e.message}",
                service_name="ProjectService"
            ) from e
        except RepositoryError as e:
            raise ServiceError(
                message=f"Error saving project to repository: {e.message}",
                service_name="ProjectService"
            ) from e
        except Exception as e:
            raise ServiceError(
                message=f"Unexpected error when saving project: {str(e)}",
                service_name="ProjectService"
            ) from e
    
    def get_project(self, project_id: str) -> Project:
        try:
            project = self._repository.get_by_id(project_id)
            
            if project is None:
                raise ServiceError(
                    message=f"No project found with ID: {project_id}",
                    service_name="ProjectService"
                )
            
            return project
            
        except RepositoryError as e:
            raise ServiceError(
                message=f"Error retrieving project from repository: {e.message}",
                service_name="ProjectService"
            ) from e
        except Exception as e:
            raise ServiceError(
                message=f"Unexpected error when retrieving project: {str(e)}",
                service_name="ProjectService"
            ) from e
    
    def get_all_projects(self) -> List[Project]:
        try:
            projects = self._repository.get_all()
            
            return projects
            
        except RepositoryError as e:
            raise ServiceError(
                message=f"Error retrieving projects from repository: {e.message}",
                service_name="ProjectService"
            ) from e
        except Exception as e:
            raise ServiceError(
                message=f"Unexpected error when retrieving projects: {str(e)}",
                service_name="ProjectService"
            ) from e
    
    def delete_project(self, project_id: str) -> bool:
        try:
            deleted = self._repository.delete(project_id)
            
            return deleted
            
        except RepositoryError as e:
            raise ServiceError(
                message=f"Error deleting project from repository: {e.message}",
                service_name="ProjectService"
            ) from e
        except Exception as e:
            raise ServiceError(
                message=f"Unexpected error when deleting project: {str(e)}",
                service_name="ProjectService"
            ) from e
    
    def search_projects(self, query: str) -> List[Project]:
        try:
            projects = self._repository.search(query)
            
            return projects
            
        except RepositoryError as e:
            raise ServiceError(
                message=f"Error searching projects in repository: {e.message}",
                service_name="ProjectService"
            ) from e
        except Exception as e:
            raise ServiceError(
                message=f"Unexpected error when searching projects: {str(e)}",
                service_name="ProjectService"
            ) from e
    
    def duplicate_project(self, project_id: str, new_name: Optional[str] = None) -> Project:
        try:
            original = self.get_project(project_id)
            
            project_dict = original.to_dict()
        
            if 'id' in project_dict:
                project_dict.pop('id')
            
            if new_name:
                project_dict['name'] = new_name
            else:
                project_dict['name'] = f"Copy of {project_dict['name']}"
            
            new_project = Project.from_dict(project_dict)
            
            # Asegurarse de que el nombre se establece correctamente
            new_project.name = project_dict['name']
            
            saved_project = self.save_project(new_project)
            
            return saved_project
            
        except ServiceError as e:
            raise ServiceError(
                message=f"Error duplicating project: {e.message}",
                service_name="ProjectService"
            ) from e
        except Exception as e:
            raise ServiceError(
                message=f"Unexpected error when duplicating project: {str(e)}",
                service_name="ProjectService"
            ) from e
    
    def export_to_excel(self, project: Project, file_path: str) -> None:
        try:
            wb = Workbook()

            ws_info = wb.active
            ws_info.title = "Project Information"

            ws_info.append(["ID", project.id])
            ws_info.append(["Name", project.name])
            ws_info.append(["Description", project.description])
            ws_info.append(["Decision Maker", project.decision_maker])
            ws_info.append(["Creation Date", project.created_at.isoformat()])
            ws_info.append(["Update Date", project.updated_at.isoformat()])

            if project.alternatives:
                ws_alternatives = wb.create_sheet('Alternatives')

                ws_alternatives.append(['ID', 'Name', 'Description'])

                for alt in project.alternatives:
                    ws_alternatives.append([alt.id, alt.name, alt.description])
            
            if project.criteria:
                ws_criteria = wb.create_sheet("Criteria")
                
                ws_criteria.append(["ID","Name","Description","Optimization Type", "Scale Type", "Weight","Unit"])

                for crit in project.criteria:
                    ws_criteria.append([
                    crit.id, 
                    crit.name, 
                    crit.description, 
                    crit.optimization_type.value,
                    crit.scale_type.value,
                    crit.weight,
                    crit.unit
                ])
            
            if project.decision_matrix is not None:
                matrix = project.decision_matrix
                ws_matrix = wb.create_sheet("Decision Matrix")

                headers = ["Alternative"] + [crit.name for crit in matrix.criteria]
                ws_matrix.append(headers)

                for i, alt in enumerate(matrix.alternative):
                    row = [alt.name]
                    for j in range(len(matrix.criteria)):
                        row.append(float(matrix.values[i, j]))
                    ws_matrix.append(row)
                
            for method_name, result in project.results.items():
                ws_result = wb.create_sheet(f"Result {method_name}")

                ws_result.append(["Method", result.method_name])
                ws_result.append(["Execution Time", f"{result.execution_time:.4f} seconds"])
                ws_result.append(["Date", result.created_at.isoformat()])
                ws_result.append([])

                ws_result.append(["ID","Name","Score","Ranking"])

                sorted_results = result.get_sorted_alternatives()

                for alt_result in sorted_results:
                    ws_result.append([
                        alt_result['id'],
                        alt_result['name'],
                        alt_result['score'],
                        alt_result['ranking']
                    ])
            wb.save(file_path)
        except Exception as e:
            raise ServiceError(
                message=f"Error exporting project to Excel: {str(e)}",
                service_name="ProjectService"
            ) from e

    def export_to_csv(self, project: Project, file_path: str) -> None:
        try:
            dir_path = os.path.dirname(file_path)
            base_name = os.path.splitext(os.path.basename(file_path))[0]

            os.makedirs(dir_path, exist_ok=True)

            info_path = os.path.join(dir_path, f"{base_name}_info.csv")
            info_path = os.path.join(dir_path, f"{base_name}_info.csv")
            with open(info_path, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["Field", "Value"])
                writer.writerow(["ID", project.id])
                writer.writerow(["Name", project.name])
                writer.writerow(["Description", project.description])
                writer.writerow(["Decision Maker", project.decision_maker])
                writer.writerow(["Creation Date", project.created_at.isoformat()])
                writer.writerow(["Update Date", project.updated_at.isoformat()])
            
            if project.alternatives:
                alt_path = os.path.join(dir_path, f"{base_name}_alternatives.csv")
                with open(alt_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(["ID", "Name", "Description"])
                    for alt in project.alternatives:
                        writer.writerow([alt.id, alt.name, alt.description])
            
            if project.criteria:
                crit_path = os.path.join(dir_path, f"{base_name}_criteria.csv")
                with open(crit_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    writer.writerow(["ID", "Name", "Description", "Optimization Type", 
                                    "Scale Type", "Weight", "Unit"])
                    for crit in project.criteria:
                        writer.writerow([
                            crit.id, 
                            crit.name, 
                            crit.description, 
                            crit.optimization_type.value,
                            crit.scale_type.value,
                            crit.weight,
                            crit.unit
                        ])
            
            if project.decision_matrix is not None:
                matrix = project.decision_matrix
                matrix_path = os.path.join(dir_path, f"{base_name}_matrix.csv")
                with open(matrix_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    
                    # Headers (criteria names)
                    headers = ["Alternative"] + [crit.name for crit in matrix.criteria]
                    writer.writerow(headers)
                    
                    # Data
                    for i, alt in enumerate(matrix.alternative):
                        row = [alt.name]
                        for j in range(len(matrix.criteria)):
                            row.append(float(matrix.values[i, j]))
                        writer.writerow(row)
            
            for method_name, result in project.results.items():
                result_path = os.path.join(dir_path, f"{base_name}_result_{method_name}.csv")
                with open(result_path, 'w', newline='', encoding='utf-8') as f:
                    writer = csv.writer(f)
                    
                    # General result information
                    writer.writerow(["Method", result.method_name])
                    writer.writerow(["Execution Time", f"{result.execution_time:.4f} seconds"])
                    writer.writerow(["Date", result.created_at.isoformat()])
                    writer.writerow([])
                    
                    # Results by alternative
                    writer.writerow(["ID", "Name", "Score", "Ranking"])
                    
                    # Sort by score (highest to lowest)
                    sorted_results = result.get_sorted_alternatives()
                    
                    for alt_result in sorted_results:
                        writer.writerow([
                            alt_result['id'],
                            alt_result['name'],
                            alt_result['score'],
                            alt_result['ranking']
                        ])
                    
        except Exception as e:
            # Convert exceptions to ServiceError
            raise ServiceError(
                message=f"Error exporting project to CSV: {str(e)}",
                service_name="ProjectService"
            ) from e

    def export_to_pdf(self, project: Project, file_path: str) -> None:
        try:
            doc = SimpleDocTemplate(file_path, pagesize=letter)
            elements = []

            styles = getSampleStyleSheet()
            title_style = styles["Heading1"]
            subtitle_style = styles['Heading2']
            normal_style = styles['Normal']

            elements.append(Paragraph("General Information", subtitle_style))
            elements.append(Spacer(1,6))

            info_data = [
                ["ID:", project.id],
                ["Description:", project.description],
                ["Decision Maker:", project.decision_maker],
                ["Creation Date:", project.created_at.isoformat()],
                ["Update Date:", project.updated_at.isoformat()]
            ]

            info_table = Table(info_data, colWidths=[120, 350])
            info_table.setStyle(TableStyle([
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                ('VALIGN', (0, 0), (-1, -1), 'MIDDLE')
            ]))

            elements.append(info_table)
            elements.append(Spacer(1, 12))

            if project.alternatives:
                elements.append(Paragraph("Alternatives", subtitle_style))
                elements.append(Spacer(1, 6))

                alt_data = [["ID", "Name", 'Description']]

                for alt in project.alternatives:
                    alt_data.append([alt.id, alt.name, alt.description])
                
                alt_table = Table(alt_data, colWidths=[80, 150, 240])
                alt_table.setStyle(TableStyle([
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))

                elements.append(alt_table)
                elements.append(Spacer(1, 12))
            
            if project.criteria:
                elements.append(Paragraph("Criteria", subtitle_style))
                elements.append(Spacer(1, 6))
                
                crit_data = [["ID", "Name", "Type", "Weight", "Unit"]]
                
                for crit in project.criteria:
                    crit_data.append([
                        crit.id, 
                        crit.name, 
                        crit.optimization_type.value,
                        str(crit.weight),
                        crit.unit
                    ])
                    
                crit_table = Table(crit_data, colWidths=[60, 150, 100, 60, 100])
                crit_table.setStyle(TableStyle([
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                
                elements.append(crit_table)
                elements.append(Spacer(1, 12))
            
            if project.decision_matrix is not None:
                matrix = project.decision_matrix
                elements.append(Paragraph("Decision Matrix", subtitle_style))
                elements.append(Spacer(1, 6))
                
                # Headers (criteria names)
                headers = ["Alternative"] + [crit.name for crit in matrix.criteria]
                matrix_data = [headers]
                
                # Data
                for i, alt in enumerate(matrix.alternative):
                    row = [alt.name]
                    for j in range(len(matrix.criteria)):
                        row.append(f"{float(matrix.values[i, j]):.4f}")
                    matrix_data.append(row)
                    
                matrix_table = Table(matrix_data)
                matrix_table.setStyle(TableStyle([
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                
                elements.append(matrix_table)
                elements.append(Spacer(1, 12))
            
            # Results
            for method_name, result in project.results.items():
                elements.append(Paragraph(f"Result: {method_name}", subtitle_style))
                elements.append(Spacer(1, 6))
                
                # Method information
                method_info = [
                    ["Method:", result.method_name],
                    ["Execution Time:", f"{result.execution_time:.4f} seconds"],
                    ["Date:", result.created_at.isoformat()]
                ]
                
                method_table = Table(method_info, colWidths=[120, 350])
                method_table.setStyle(TableStyle([
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('BACKGROUND', (0, 0), (0, -1), colors.lightgrey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                
                elements.append(method_table)
                elements.append(Spacer(1, 6))
                
                # Results by alternative
                result_data = [["Ranking", "ID", "Name", "Score"]]
                
                # Sort by score (highest to lowest)
                sorted_results = result.get_sorted_alternatives()
                
                for alt_result in sorted_results:
                    result_data.append([
                        str(alt_result['ranking']),
                        alt_result['id'],
                        alt_result['name'],
                        f"{alt_result['score']:.4f}"
                    ])
                    
                result_table = Table(result_data, colWidths=[60, 80, 200, 130])
                result_table.setStyle(TableStyle([
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                    ('BACKGROUND', (0, 0), (-1, 0), colors.lightgrey),
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                ]))
                
                elements.append(result_table)
                elements.append(Spacer(1, 12))
            
            # Generate PDF
            doc.build(elements)
                
        except Exception as e:
            # Convert exceptions to ServiceError
            raise ServiceError(
                message=f"Error exporting project to PDF: {str(e)}",
                service_name="ProjectService"
            ) from e

    def export_to_json(self, project: Project, file_path: str) -> None:
        try:
            project_dict = project.to_dict()
            
            os.makedirs(os.path.dirname(os.path.abspath(file_path)), exist_ok=True)
            
            with open(file_path, 'w', encoding='utf-8') as f:
                json.dump(project_dict, f, indent=2, ensure_ascii=False)
                
        except Exception as e:
            raise ServiceError(
                message=f"Error exporting project to JSON: {str(e)}",
                service_name="ProjectService"
            ) from e
    
    def import_from_json(self, file_path: str) -> Project:
        try:
            with open(file_path, 'r', encoding='utf-8') as f:
                project_dict = json.load(f)
            
            if not isinstance(project_dict, dict):
                raise ValidationError(
                    message="The JSON file does not contain an object",
                    errors=["The file must contain a valid JSON object representing a project"]
                )
            
            valid, errors = ProjectValidator.validate_from_dict(project_dict)
            if not valid:
                raise ValidationError(
                    message="The JSON file does not contain a valid project",
                    errors=errors
                )
            
            project = Project.from_dict(project_dict)
            
            return project
            
        except ValidationError as e:
            raise ServiceError(
                message=f"Validation error when importing project: {e.message}",
                service_name="ProjectService"
            ) from e
        except json.JSONDecodeError as e:
            raise ServiceError(
                message=f"Error decoding JSON: {str(e)}",
                service_name="ProjectService"
            ) from e
        except Exception as e:
            raise ServiceError(
                message=f"Unexpected error when importing project: {str(e)}",
                service_name="ProjectService"
            ) from e
    
    def import_from_excel(self, file_path: str) -> Project:
        try:
            import pandas as pd
            from domain.entities.project import Project
            from domain.entities.alternative import Alternative
            from domain.entities.criteria import Criteria, OptimizationType, ScaleType
            
            xls = pd.ExcelFile(file_path)
            
            # Verify that the necessary sheets exist
            required_sheets = ["Project Information", "Alternatives", "Criteria"]
            for sheet in required_sheets:
                if sheet not in xls.sheet_names:
                    raise ValidationError(
                        message=f"Invalid Excel file format",
                        errors=[f"Sheet '{sheet}' not found"]
                    )
            
            df_info = pd.read_excel(xls, "Project Information", header=None)
            
            info_dict = {}
            for _, row in df_info.iterrows():
                if len(row) >= 2:
                    info_dict[row[0]] = row[1]
            
            project = Project(
                name=info_dict.get("Name", "Imported Project"),
                description=info_dict.get("Description", ""),
                decision_maker=info_dict.get("Decision Maker", ""),
                project_id=info_dict.get("ID")
            )
            
            df_alternatives = pd.read_excel(xls, "Alternatives")
            
            for _, row in df_alternatives.iterrows():
                alternative = Alternative(
                    id=row["ID"],
                    name=row["Name"],
                    description=row.get("Description", "")
                )
                project.add_alternative(alternative)
            
            df_criteria = pd.read_excel(xls, "Criteria")
            
            for _, row in df_criteria.iterrows():
                opt_type = row.get("Optimization Type", "maximize")
                scale_type = row.get("Scale Type", "quantitative")
                
                if isinstance(opt_type, str):
                    opt_type = OptimizationType(opt_type)
                    
                if isinstance(scale_type, str):
                    scale_type = ScaleType(scale_type)
                
                criteria = Criteria(
                    id=row["ID"],
                    name=row["Name"],
                    description=row.get("Description", ""),
                    optimization_type=opt_type,
                    scale_type=scale_type,
                    weight=float(row.get("Weight", 1.0)),
                    unit=row.get("Unit", "")
                )
                project.add_criteria(criteria)
            
            # If the matrix sheet exists, create decision matrix
            if "Decision Matrix" in xls.sheet_names:
                df_matrix = pd.read_excel(xls, "Decision Matrix")
                
                # Convert DataFrame to value matrix
                alt_names = project.alternatives
                crit_names = project.criteria
                
                if len(alt_names) > 0 and len(crit_names) > 0:
                    # Create empty matrix
                    matrix = project.create_decision_matrix()
                    
                    # Fill values
                    for i, alt in enumerate(alt_names):
                        for j, crit in enumerate(crit_names):
                            try:
                                # Search for value in DataFrame
                                alt_row = df_matrix[df_matrix.iloc[:, 0] == alt.name]
                                if not alt_row.empty:
                                    value = float(alt_row.iloc[0, j + 1])
                                    matrix.set_values(i, j, value)
                            except (IndexError, ValueError):
                                # If there's an error, leave default value (0)
                                pass
            
            return project
                
        except ValidationError as e:
            raise ServiceError(
                message=f"Validation error when importing project from Excel: {e.message}",
                service_name="ProjectService"
            ) from e
        except Exception as e:
            raise ServiceError(
                message=f"Error importing project from Excel: {str(e)}",
                service_name="ProjectService"
            ) from e

    def import_from_csv(self, file_path: str) -> Project:
        try:
            
            
            # Extract directory and base name of the file
            dir_path = os.path.dirname(file_path)
            base_name = os.path.splitext(os.path.basename(file_path))[0]
            
            # Verify that the information file exists
            info_path = os.path.join(dir_path, f"{base_name}_info.csv")
            if not os.path.exists(info_path):
                # Search for any file with suffix _info.csv
                info_files = [f for f in os.listdir(dir_path) if f.endswith("_info.csv")]
                if info_files:
                    info_path = os.path.join(dir_path, info_files[0])
                    base_name = info_files[0].replace("_info.csv", "")
                else:
                    raise ValidationError(
                        message="Information file not found",
                        errors=["You must provide a file with suffix _info.csv"]
                    )
            
            # Read project information
            info_dict = {}
            with open(info_path, 'r', newline='', encoding='utf-8') as f:
                reader = csv.reader(f)
                next(reader)  # Skip headers
                for row in reader:
                    if len(row) >= 2:
                        info_dict[row[0]] = row[1]
            
            # Create project
            project = Project(
                name=info_dict.get("Name", "Imported Project"),
                description=info_dict.get("Description", ""),
                decision_maker=info_dict.get("Decision Maker", ""),
                project_id=info_dict.get("ID")
            )
            
            # Read alternatives
            alt_path = os.path.join(dir_path, f"{base_name}_alternatives.csv")
            if os.path.exists(alt_path):
                with open(alt_path, 'r', newline='', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    headers = next(reader)  # Read headers
                    
                    for row in reader:
                        if len(row) >= 2:
                            alternative = Alternative(
                                id=row[0],
                                name=row[1],
                                description=row[2] if len(row) > 2 else ""
                            )
                            project.add_alternative(alternative)
            
            # Read criteria
            crit_path = os.path.join(dir_path, f"{base_name}_criteria.csv")
            if os.path.exists(crit_path):
                with open(crit_path, 'r', newline='', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    headers = next(reader)  # Read headers
                    
                    # Crear un mapeo de nombres de columna a índices
                    col_map = {header: idx for idx, header in enumerate(headers)}
                    
                    for row in reader:
                        if len(row) >= 2:  # Al menos ID y Name
                            # Usar el mapeo de columnas para obtener los valores correctos
                            id_val = row[col_map.get('ID', 0)]
                            name_val = row[col_map.get('Name', 1)]
                            desc_val = row[col_map.get('Description', 2)] if 'Description' in col_map and len(row) > col_map['Description'] else ""
                            
                            # Obtener optimization type
                            opt_type_idx = col_map.get('Optimization Type', -1)
                            opt_type = row[opt_type_idx] if opt_type_idx >= 0 and len(row) > opt_type_idx else "maximize"
                            
                            # Obtener scale type
                            scale_type_idx = col_map.get('Scale Type', -1)
                            scale_type = row[scale_type_idx] if scale_type_idx >= 0 and len(row) > scale_type_idx else "quantitative"
                            
                            # Obtener weight
                            weight_idx = col_map.get('Weight', -1)
                            weight = float(row[weight_idx]) if weight_idx >= 0 and len(row) > weight_idx else 1.0
                            
                            # Obtener unit
                            unit_idx = col_map.get('Unit', -1)
                            unit = row[unit_idx] if unit_idx >= 0 and len(row) > unit_idx else ""
                            
                            # Convertir a enums
                            if isinstance(opt_type, str):
                                opt_type = OptimizationType(opt_type)
                                
                            if isinstance(scale_type, str):
                                scale_type = ScaleType(scale_type)
                            
                            criteria = Criteria(
                                id=id_val,
                                name=name_val,
                                description=desc_val,
                                optimization_type=opt_type,
                                scale_type=scale_type,
                                weight=weight,
                                unit=unit
                            )
                            project.add_criteria(criteria)
            
            # Read decision matrix
            matrix_path = os.path.join(dir_path, f"{base_name}_matrix.csv")
            if os.path.exists(matrix_path) and project.alternatives and project.criteria:
                matrix = project.create_decision_matrix()
                
                with open(matrix_path, 'r', newline='', encoding='utf-8') as f:
                    reader = csv.reader(f)
                    headers = next(reader)  

                    for i, row in enumerate(reader):
                        if len(row) > 1:
                            alt_name = row[0]
                            
                            alt_idx = None
                            for j, alt in enumerate(project.alternatives):
                                if alt.name == alt_name:
                                    alt_idx = j
                                    break
                            
                            if alt_idx is not None:
                                for j in range(1, min(len(row), len(project.criteria) + 1)):
                                    try:
                                        value = float(row[j])
                                        matrix.set_values(alt_idx, j - 1, value)
                                    except (ValueError, IndexError):
                                        
                                        pass
            
            return project
                
        except ValidationError as e:
            raise ServiceError(
                message=f"Validation error when importing project from CSV: {e.message}",
                service_name="ProjectService"
            ) from e
        except Exception as e:
            raise ServiceError(
                message=f"Error importing project from CSV: {str(e)}",
                service_name="ProjectService"
            ) from e